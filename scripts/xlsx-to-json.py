#!/usr/bin/env python3
"""Converte o export xlsx do Jestor em JSON bruto para o importador.

Extração pura: nenhuma regra de negócio aqui. Toda normalização e mapeamento
ficam em scripts/src/import-jestor.ts, para haver uma única fonte de verdade.

Uso:
    python scripts/xlsx-to-json.py <arquivo.xlsx> <saida.json>

O JSON de saída contém dados de clientes — não versionar.
"""
import json
import sys

import openpyxl


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 1

    src, dest = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    rows = []
    for excel_row in ws.iter_rows(min_row=2):
        values = [c.value for c in excel_row]
        if all(v in (None, "") for v in values):
            continue
        row = {}
        for header, value in zip(headers, values):
            if header is None:
                continue
            # openpyxl devolve datetime para algumas células; serializa como texto.
            row[header] = value if isinstance(value, (str, int, float, type(None))) else str(value)
        rows.append(row)

    with open(dest, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, ensure_ascii=False, indent=2)

    print(f"{len(rows)} linhas -> {dest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
